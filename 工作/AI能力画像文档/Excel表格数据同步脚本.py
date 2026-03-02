import openpyxl
from pathlib import Path

def sync_excel_sheets(file_path):
    """
    对比并同步 Excel 文件中 Sheet1 和 Sheet2 的数据
    
    参数:
        file_path: Excel 文件路径
    """
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    
    # 获取两个工作表
    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']
    
    # 构建 Sheet1 的 A 列索引字典（值 -> 行号）
    # 从第2行开始读取
    sheet1_a_column = {}
    for row in range(2, sheet1.max_row + 1):
        cell_value = sheet1.cell(row=row, column=1).value
        if cell_value is not None:
            sheet1_a_column[cell_value] = row
    
    # 遍历 Sheet2 的数据（从第2行开始）
    for row in range(2, sheet2.max_row + 1):
        # 获取 Sheet2 当前行的 A 列值
        a_value = sheet2.cell(row=row, column=1).value
        
        # 如果这个值在 Sheet1 中存在
        if a_value in sheet1_a_column:
            target_row = sheet1_a_column[a_value]
            
            # 复制 Sheet2 当前行的所有数据到 Sheet1 对应行的 L 列开始的位置
            # L 列是第 12 列
            col_offset = 12  # L 列的列号
            
            # 获取 Sheet2 当前行的最大列数
            for col in range(1, sheet2.max_column + 1):
                source_cell = sheet2.cell(row=row, column=col)
                target_cell = sheet1.cell(row=target_row, column=col_offset + col - 1)
                
                # 复制单元格的值
                target_cell.value = source_cell.value
                
                # 复制单元格的格式（可选）
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()
    
    # 保存工作簿
    wb.save(file_path)
    print(f"处理完成！文件已保存：{file_path}")
    print(f"Sheet1 中找到 {len(sheet1_a_column)} 个 A 列数据")
    print(f"Sheet2 中处理了 {sheet2.max_row - 1} 行数据")

# 执行脚本
if __name__ == "__main__":
    file_path = r"D:\proiect\工作\AI能力画像文档\同步更新.xlsx"
    
    # 检查文件是否存在
    if not Path(file_path).exists():
        print(f"错误：文件不存在 - {file_path}")
    else:
        try:
            sync_excel_sheets(file_path)
        except Exception as e:
            print(f"处理过程中出现错误：{e}")
            import traceback
            traceback.print_exc()