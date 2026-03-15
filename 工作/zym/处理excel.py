import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import threading
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ---------- 排班逻辑 ----------

WORK_VALUES = [x / 2 for x in range(12, 23)]  # 6.0, 6.5, … 11.0

# 排班模式: (工作天数, 休息天数)
SHIFT_PATTERNS = [(6, 1), (5, 1), (5, 2)]


def generate_column_data(num_rows: int) -> list:
    """
    为单列生成随机排班数据。
    - 第一天允许直接休息（30% 概率）
    - 同一列内随机组合多种排班模式
    - 工作日填随机数值（6–11，步长 0.5），休息日填 None
    """
    result = []

    # 决定第一天是否休息
    if random.random() < 0.3:
        result.append(None)

    while len(result) < num_rows:
        work, rest = random.choice(SHIFT_PATTERNS)

        for _ in range(work):
            if len(result) >= num_rows:
                break
            result.append(random.choice(WORK_VALUES))

        for _ in range(rest):
            if len(result) >= num_rows:
                break
            result.append(None)

    return result[:num_rows]


def write_excel(
    filepath: str,
    col_start: str,
    col_end: str,
    row_start: int,
    row_end: int,
    progress_cb=None,
):
    """将随机排班数据写入 Excel 文件。"""
    ci_start = column_index_from_string(col_start.upper())
    ci_end = column_index_from_string(col_end.upper())
    if ci_start > ci_end:
        ci_start, ci_end = ci_end, ci_start
    if row_start > row_end:
        row_start, row_end = row_end, row_start

    num_rows = row_end - row_start + 1
    total_cols = ci_end - ci_start + 1

    # 尝试打开已有文件，否则新建
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "排班数据"

    # 样式定义
    work_fill = PatternFill("solid", fgColor="DDEEFF")
    rest_fill = PatternFill("solid", fgColor="FFE4E4")
    work_font = Font(name="Arial", size=10, color="1A3C5E")
    rest_font = Font(name="Arial", size=10, color="CC0000", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, ci in enumerate(range(ci_start, ci_end + 1)):
        col_letter = get_column_letter(ci)
        data = generate_column_data(num_rows)

        for row_offset, value in enumerate(data):
            cell = ws.cell(row=row_start + row_offset, column=ci)
            cell.alignment = center
            cell.border = border

            if value is None:
                cell.value = "休"
                cell.font = rest_font
                cell.fill = rest_fill
            else:
                cell.value = value
                cell.font = work_font
                cell.fill = work_fill

        ws.column_dimensions[col_letter].width = 8

        if progress_cb:
            progress_cb(idx + 1, total_cols)

    wb.save(filepath)


# ---------- GUI ----------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 排班数据生成器")
        self.resizable(False, False)
        self._build_ui()

    # ---- UI 构建 ----

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        # ── 文件选择 ──────────────────────────────────────────
        file_frame = ttk.LabelFrame(self, text="输出文件", padding=8)
        file_frame.grid(row=0, column=0, columnspan=2, sticky="ew", **pad)

        self.filepath_var = tk.StringVar(value="排班数据.xlsx")
        ttk.Entry(file_frame, textvariable=self.filepath_var, width=42).grid(
            row=0, column=0, padx=(0, 6)
        )
        ttk.Button(file_frame, text="浏览…", command=self._browse).grid(
            row=0, column=1
        )

        # ── 列范围 ────────────────────────────────────────────
        range_frame = ttk.LabelFrame(self, text="数据范围", padding=8)
        range_frame.grid(row=1, column=0, columnspan=2, sticky="ew", **pad)

        labels = ["起始列", "结束列", "起始行", "结束行"]
        defaults = ["C", "AZ", "5", "46"]
        self.range_vars = []

        for i, (lbl, dflt) in enumerate(zip(labels, defaults)):
            ttk.Label(range_frame, text=lbl + "：").grid(
                row=i // 2, column=(i % 2) * 2, sticky="e", padx=(0, 4), pady=3
            )
            var = tk.StringVar(value=dflt)
            ttk.Entry(range_frame, textvariable=var, width=8).grid(
                row=i // 2, column=(i % 2) * 2 + 1, sticky="w", padx=(0, 16), pady=3
            )
            self.range_vars.append(var)

        # ── 进度 ──────────────────────────────────────────────
        prog_frame = ttk.Frame(self, padding=(10, 0))
        prog_frame.grid(row=2, column=0, columnspan=2, sticky="ew")

        self.progress = ttk.Progressbar(prog_frame, length=380, mode="determinate")
        self.progress.grid(row=0, column=0, sticky="ew")
        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(prog_frame, textvariable=self.status_var, width=48).grid(
            row=1, column=0, sticky="w", pady=(2, 0)
        )

        # ── 按钮 ──────────────────────────────────────────────
        btn_frame = ttk.Frame(self, padding=(10, 4))
        btn_frame.grid(row=3, column=0, columnspan=2, sticky="e")

        self.gen_btn = ttk.Button(
            btn_frame, text="生成数据", command=self._start_generate, width=14
        )
        self.gen_btn.grid(row=0, column=0, padx=4)
        ttk.Button(btn_frame, text="退出", command=self.destroy, width=8).grid(
            row=0, column=1
        )

        # ── 说明 ──────────────────────────────────────────────
        info = (
            "排班模式：上六休一 / 上五休一 / 上五休二（列内随机混合）\n"
            "数值范围：6 ~ 11（步长 0.5）；休息日显示「休」"
        )
        ttk.Label(
            self, text=info, foreground="#555", font=("Arial", 8), justify="left"
        ).grid(row=4, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 8))

    # ---- 事件处理 ----

    def _browse(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=self.filepath_var.get(),
        )
        if path:
            self.filepath_var.set(path)

    def _validate_inputs(self):
        col_start = self.range_vars[0].get().strip()
        col_end = self.range_vars[1].get().strip()
        row_start_s = self.range_vars[2].get().strip()
        row_end_s = self.range_vars[3].get().strip()
        filepath = self.filepath_var.get().strip()

        if not filepath:
            messagebox.showerror("错误", "请指定输出文件路径。")
            return None

        if not col_start.isalpha() or not col_end.isalpha():
            messagebox.showerror("错误", "列标识符只能包含字母（如 C、AZ）。")
            return None

        try:
            row_start = int(row_start_s)
            row_end = int(row_end_s)
        except ValueError:
            messagebox.showerror("错误", "行号必须为整数。")
            return None

        if row_start < 1 or row_end < 1:
            messagebox.showerror("错误", "行号必须 ≥ 1。")
            return None

        return filepath, col_start, col_end, row_start, row_end

    def _start_generate(self):
        params = self._validate_inputs()
        if not params:
            return

        self.gen_btn.configure(state="disabled")
        self.progress["value"] = 0
        self.status_var.set("正在生成…")

        def run():
            filepath, col_start, col_end, row_start, row_end = params
            ci_start = column_index_from_string(col_start.upper())
            ci_end = column_index_from_string(col_end.upper())
            total = abs(ci_end - ci_start) + 1

            def on_progress(done, total_cols):
                pct = done / total_cols * 100
                self.progress["value"] = pct
                self.status_var.set(f"已处理 {done}/{total_cols} 列…")

            try:
                write_excel(filepath, col_start, col_end, row_start, row_end, on_progress)
                self.after(
                    0,
                    lambda: messagebox.showinfo(
                        "完成",
                        f"数据生成完毕！\n共写入 {total} 列 × {abs(row_end - row_start) + 1} 行\n→ {filepath}",
                    ),
                )
                self.status_var.set("生成完成 ✓")
            except Exception as e:
                self.after(
                    0, lambda: messagebox.showerror("错误", f"生成失败：{e}")
                )
                self.status_var.set("生成失败 ✗")
            finally:
                self.after(0, lambda: self.gen_btn.configure(state="normal"))

        threading.Thread(target=run, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()