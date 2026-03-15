import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import threading
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ---------- 排班逻辑 ----------

WORK_VALUES = [x / 2 for x in range(12, 23)]  # 6.0, 6.5, … 11.0
SHIFT_PATTERNS = [(6, 1), (5, 1), (5, 2)]      # (工作天数, 休息天数)


def generate_column_data(num_rows: int) -> list:
    """
    为单列生成随机排班数据。
    - 第一天允许直接休息（30% 概率）
    - 同一列内随机组合多种排班模式
    - 工作日填随机数值（6–11，步长 0.5），休息日填 None
    """
    result = []
    if random.random() < 0.3:
        result.append(None)

    while len(result) < num_rows:
        work, rest = random.choice(SHIFT_PATTERNS)
        for _ in range(work):
            if len(result) >= num_rows:
                break
            result.append(random.choice(WORK_VALUES))
        for _ in range(rest):
            if len(result) >= num_rows:
                break
            result.append(None)

    return result[:num_rows]


def write_excel(input_path, output_path, col_start, col_end,
                row_start, row_end, progress_cb=None):
    """读取输入 Excel，将随机排班数据写入指定范围，保存到输出路径。"""
    ci_start = column_index_from_string(col_start.upper())
    ci_end   = column_index_from_string(col_end.upper())
    if ci_start > ci_end:
        ci_start, ci_end = ci_end, ci_start
    if row_start > row_end:
        row_start, row_end = row_end, row_start

    num_rows   = row_end - row_start + 1
    total_cols = ci_end - ci_start + 1

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 样式定义
    work_fill = PatternFill("solid", fgColor="DDEEFF")
    rest_fill = PatternFill("solid", fgColor="FFE4E4")
    work_font = Font(name="Arial", size=10, color="1A3C5E")
    rest_font = Font(name="Arial", size=10, color="CC0000", bold=True)
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, ci in enumerate(range(ci_start, ci_end + 1)):
        col_letter = get_column_letter(ci)
        data = generate_column_data(num_rows)

        for row_offset, value in enumerate(data):
            cell = ws.cell(row=row_start + row_offset, column=ci)
            cell.alignment = center
            cell.border    = border
            if value is None:
                cell.value = "休"
                cell.font  = rest_font
                cell.fill  = rest_fill
            else:
                cell.value = value
                cell.font  = work_font
                cell.fill  = work_fill

        ws.column_dimensions[col_letter].width = 8

        if progress_cb:
            progress_cb(idx + 1, total_cols)

    wb.save(output_path)


# ---------- GUI ----------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 排班数据生成器")
        self.resizable(False, False)
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        # ── 输入文件 ──────────────────────────────────────────
        file_frame = ttk.LabelFrame(self, text="输入文件", padding=8)
        file_frame.grid(row=0, column=0, columnspan=2, sticky="ew", **pad)

        self.filepath_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.filepath_var,
                  width=40, state="readonly").grid(row=0, column=0, padx=(0, 6))
        ttk.Button(file_frame, text="选择文件…", command=self._browse_input).grid(
            row=0, column=1)

        # ── 处理范围 ──────────────────────────────────────────
        range_frame = ttk.LabelFrame(self, text="处理范围（可自定义）", padding=8)
        range_frame.grid(row=1, column=0, columnspan=2, sticky="ew", **pad)

        labels   = ["起始列", "结束列", "起始行", "结束行"]
        defaults = ["C",     "AZ",    "5",     "46"]
        self.range_vars = []

        for i, (lbl, dflt) in enumerate(zip(labels, defaults)):
            ttk.Label(range_frame, text=lbl + "：").grid(
                row=i // 2, column=(i % 2) * 2,
                sticky="e", padx=(0, 4), pady=3)
            var = tk.StringVar(value=dflt)
            ttk.Entry(range_frame, textvariable=var, width=8).grid(
                row=i // 2, column=(i % 2) * 2 + 1,
                sticky="w", padx=(0, 20), pady=3)
            self.range_vars.append(var)

        # ── 保存方式 ──────────────────────────────────────────
        save_frame = ttk.LabelFrame(self, text="保存方式", padding=8)
        save_frame.grid(row=2, column=0, columnspan=2, sticky="ew", **pad)

        self.save_mode = tk.StringVar(value="overwrite")
        ttk.Radiobutton(save_frame, text="覆盖原文件",
                        variable=self.save_mode, value="overwrite",
                        command=self._toggle_saveas).grid(
            row=0, column=0, padx=(0, 16), sticky="w")
        ttk.Radiobutton(save_frame, text="另存为新文件",
                        variable=self.save_mode, value="saveas",
                        command=self._toggle_saveas).grid(
            row=0, column=1, sticky="w")

        self.saveas_var = tk.StringVar()
        self.saveas_entry = ttk.Entry(save_frame, textvariable=self.saveas_var,
                                      width=30, state="disabled")
        self.saveas_entry.grid(row=1, column=0, columnspan=2,
                               pady=(6, 0), padx=(0, 6), sticky="w")
        self.saveas_btn = ttk.Button(save_frame, text="选择路径…",
                                     command=self._browse_saveas, state="disabled")
        self.saveas_btn.grid(row=1, column=2, pady=(6, 0))

        # ── 进度 ──────────────────────────────────────────────
        prog_frame = ttk.Frame(self, padding=(10, 0))
        prog_frame.grid(row=3, column=0, columnspan=2, sticky="ew")

        self.progress = ttk.Progressbar(prog_frame, length=390, mode="determinate")
        self.progress.grid(row=0, column=0, sticky="ew")
        self.status_var = tk.StringVar(value="请先选择一个 Excel 文件")
        ttk.Label(prog_frame, textvariable=self.status_var, width=50).grid(
            row=1, column=0, sticky="w", pady=(2, 0))

        # ── 按钮 ──────────────────────────────────────────────
        btn_frame = ttk.Frame(self, padding=(10, 4))
        btn_frame.grid(row=4, column=0, columnspan=2, sticky="e")

        self.gen_btn = ttk.Button(btn_frame, text="开始处理",
                                  command=self._start_generate,
                                  width=14, state="disabled")
        self.gen_btn.grid(row=0, column=0, padx=4)
        ttk.Button(btn_frame, text="退出", command=self.destroy, width=8).grid(
            row=0, column=1)

        # ── 说明 ──────────────────────────────────────────────
        info = ("排班模式：上六休一 / 上五休一 / 上五休二（列内随机混合）\n"
                "数值范围：6 ~ 11（步长 0.5）；休息日显示「休」")
        ttk.Label(self, text=info, foreground="#555",
                  font=("Arial", 8), justify="left").grid(
            row=5, column=0, columnspan=2,
            sticky="w", padx=10, pady=(0, 8))

    # ---- 事件 ----

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="选择要处理的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")])
        if path:
            self.filepath_var.set(path)
            self.status_var.set(f"已选择：{path.split('/')[-1]}")
            self.gen_btn.configure(state="normal")

    def _toggle_saveas(self):
        state = "normal" if self.save_mode.get() == "saveas" else "disabled"
        self.saveas_entry.configure(state=state)
        self.saveas_btn.configure(state=state)

    def _browse_saveas(self):
        path = filedialog.asksaveasfilename(
            title="另存为",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")])
        if path:
            self.saveas_var.set(path)

    def _validate_inputs(self):
        input_path = self.filepath_var.get().strip()
        if not input_path:
            messagebox.showerror("错误", "请先选择一个 Excel 输入文件。")
            return None

        col_start   = self.range_vars[0].get().strip()
        col_end     = self.range_vars[1].get().strip()
        row_start_s = self.range_vars[2].get().strip()
        row_end_s   = self.range_vars[3].get().strip()

        if not col_start.isalpha() or not col_end.isalpha():
            messagebox.showerror("错误", "列标识符只能包含字母（如 C、AZ）。")
            return None

        try:
            row_start = int(row_start_s)
            row_end   = int(row_end_s)
        except ValueError:
            messagebox.showerror("错误", "行号必须为整数。")
            return None

        if row_start < 1 or row_end < 1:
            messagebox.showerror("错误", "行号必须 ≥ 1。")
            return None

        if self.save_mode.get() == "saveas":
            out_path = self.saveas_var.get().strip()
            if not out_path:
                messagebox.showerror("错误", "请指定另存为的文件路径。")
                return None
        else:
            out_path = input_path

        return input_path, out_path, col_start, col_end, row_start, row_end

    def _start_generate(self):
        params = self._validate_inputs()
        if not params:
            return

        self.gen_btn.configure(state="disabled")
        self.progress["value"] = 0
        self.status_var.set("正在处理…")

        def run():
            input_path, out_path, col_start, col_end, row_start, row_end = params
            ci_start   = column_index_from_string(col_start.upper())
            ci_end     = column_index_from_string(col_end.upper())
            total_cols = abs(ci_end - ci_start) + 1
            total_rows = abs(row_end - row_start) + 1

            def on_progress(done, tc):
                self.progress["value"] = done / tc * 100
                self.status_var.set(f"已处理 {done}/{tc} 列…")

            try:
                write_excel(input_path, out_path, col_start, col_end,
                            row_start, row_end, on_progress)
                self.after(0, lambda: messagebox.showinfo(
                    "完成",
                    f"处理完毕！\n写入 {total_cols} 列 × {total_rows} 行\n→ {out_path}"))
                self.status_var.set("处理完成 ✓")
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("错误", f"处理失败：{e}"))
                self.status_var.set("处理失败 ✗")
            finally:
                self.after(0, lambda: self.gen_btn.configure(state="normal"))

        threading.Thread(target=run, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
