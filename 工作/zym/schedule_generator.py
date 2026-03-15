#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
排班数据生成器 - GUI应用程序
用于生成Excel排班数据，支持多种排班模式
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os


class ScheduleGenerator:
    """排班数据生成器核心类"""

    # 排班模式定义 (工作天数, 休息天数)
    SCHEDULE_PATTERNS = {
        '上六休一': (6, 1),
        '上五休一': (5, 1),
        '上五休二': (5, 2),
    }

    # 数据取值范围和步长
    MIN_VALUE = 6
    MAX_VALUE = 11
    STEP = 0.5

    def __init__(self):
        self.values = self._generate_valid_values()

    def _generate_valid_values(self):
        """生成所有有效的数值 [6, 11]，步长0.5"""
        values = []
        current = self.MIN_VALUE
        while current <= self.MAX_VALUE:
            values.append(current)
            current += self.STEP
        return values

    def generate_column_data(self, num_rows, start_with_rest=None):
        """
        生成单列排班数据

        Args:
            num_rows: 行数
            start_with_rest: 是否从休息日开始，None表示随机决定

        Returns:
            list: 生成的数据列表（数值或None表示休息）
        """
        if start_with_rest is None:
            start_with_rest = random.choice([True, False])

        # 随机选择排班模式
        pattern_name = random.choice(list(self.SCHEDULE_PATTERNS.keys()))
        work_days, rest_days = self.SCHEDULE_PATTERNS[pattern_name]

        data = []
        current_pos = 0

        # 如果从休息日开始
        if start_with_rest and rest_days > 0:
            rest_count = min(rest_days, num_rows - current_pos)
            for _ in range(rest_count):
                data.append(None)  # None表示休息
                current_pos += 1

        # 循环生成工作和休息周期
        while current_pos < num_rows:
            # 工作周期 - 每天随机一个数值
            work_count = min(work_days, num_rows - current_pos)
            for _ in range(work_count):
                data.append(random.choice(self.values))
                current_pos += 1

            if current_pos >= num_rows:
                break

            # 休息周期
            rest_count = min(rest_days, num_rows - current_pos)
            for _ in range(rest_count):
                data.append(None)  # None表示休息
                current_pos += 1

        return data

    def generate_all_data(self, num_cols, num_rows):
        """
        生成所有列的数据

        Args:
            num_cols: 列数
            num_rows: 行数

        Returns:
            list: 每列的数据列表组成的列表
        """
        all_data = []
        for _ in range(num_cols):
            col_data = self.generate_column_data(num_rows)
            all_data.append(col_data)
        return all_data


class ScheduleGeneratorGUI:
    """排班数据生成器 GUI 类"""

    def __init__(self, root):
        self.root = root
        self.root.title("排班数据生成器")
        self.root.geometry("600x500")
        self.root.resizable(True, True)

        # 设置样式
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TLabel', background='#f0f0f0', font=('微软雅黑', 10))
        self.style.configure('TButton', font=('微软雅黑', 10))
        self.style.configure('Header.TLabel', font=('微软雅黑', 14, 'bold'))

        self.generator = ScheduleGenerator()
        self.selected_file = None

        self._create_widgets()

    def _create_widgets(self):
        """创建GUI组件"""
        # 主容器
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 配置grid权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        # 标题
        title_label = ttk.Label(main_frame, text="排班数据生成器", style='Header.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # 文件选择区域
        ttk.Label(main_frame, text="Excel文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.file_entry = ttk.Entry(main_frame, width=40)
        self.file_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(main_frame, text="浏览...", command=self._browse_file).grid(row=1, column=2, padx=5, pady=5)

        # 分隔线
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        # 区域设置区域
        ttk.Label(main_frame, text="数据区域设置", font=('微软雅黑', 11, 'bold')).grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(5, 10))

        # 起始列和结束列
        ttk.Label(main_frame, text="起始列:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.start_col_var = tk.StringVar(value="C")
        ttk.Entry(main_frame, textvariable=self.start_col_var, width=10).grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)
        ttk.Label(main_frame, text="(如: C, AZ)").grid(row=4, column=2, sticky=tk.W, pady=5)

        ttk.Label(main_frame, text="结束列:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.end_col_var = tk.StringVar(value="AZ")
        ttk.Entry(main_frame, textvariable=self.end_col_var, width=10).grid(row=5, column=1, sticky=tk.W, padx=5, pady=5)

        # 起始行和结束行
        ttk.Label(main_frame, text="起始行:").grid(row=6, column=0, sticky=tk.W, pady=5)
        self.start_row_var = tk.StringVar(value="5")
        ttk.Entry(main_frame, textvariable=self.start_row_var, width=10).grid(row=6, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(main_frame, text="结束行:").grid(row=7, column=0, sticky=tk.W, pady=5)
        self.end_row_var = tk.StringVar(value="46")
        ttk.Entry(main_frame, textvariable=self.end_row_var, width=10).grid(row=7, column=1, sticky=tk.W, padx=5, pady=5)

        # 分隔线
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        # 数据设置区域
        ttk.Label(main_frame, text="数据设置", font=('微软雅黑', 11, 'bold')).grid(row=9, column=0, columnspan=3, sticky=tk.W, pady=(5, 10))

        # 排班模式选择
        ttk.Label(main_frame, text="排班模式:").grid(row=10, column=0, sticky=tk.W, pady=5)
        self.pattern_vars = {}
        patterns_frame = ttk.Frame(main_frame)
        patterns_frame.grid(row=10, column=1, columnspan=2, sticky=tk.W, pady=5)

        for i, pattern in enumerate(ScheduleGenerator.SCHEDULE_PATTERNS.keys()):
            var = tk.BooleanVar(value=True)
            self.pattern_vars[pattern] = var
            ttk.Checkbutton(patterns_frame, text=pattern, variable=var).pack(side=tk.LEFT, padx=5)

        # 是否允许从休息日开始
        self.allow_rest_start_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(main_frame, text="允许排班从休息日开始", variable=self.allow_rest_start_var).grid(row=11, column=1, columnspan=2, sticky=tk.W, pady=5)

        # 分隔线
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).grid(row=12, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)

        # 操作按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=13, column=0, columnspan=3, pady=20)

        ttk.Button(button_frame, text="生成并保存", command=self._generate_and_save, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="新建Excel并生成", command=self._create_new_excel, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="退出", command=self.root.quit, width=15).pack(side=tk.LEFT, padx=5)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=14, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))

    def _browse_file(self):
        """浏览选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.selected_file = file_path
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.status_var.set(f"已选择文件: {os.path.basename(file_path)}")

    def _col_to_index(self, col_str):
        """将列字母转换为索引 (A=1, B=2, ..., Z=26, AA=27, ...)"""
        col_str = col_str.upper()
        index = 0
        for char in col_str:
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index

    def _index_to_col(self, index):
        """将索引转换为列字母 (1=A, 2=B, ...)"""
        result = ""
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _get_selected_patterns(self):
        """获取用户选择的排班模式"""
        selected = []
        for pattern, var in self.pattern_vars.items():
            if var.get():
                selected.append(pattern)
        return selected

    def _validate_inputs(self):
        """验证输入参数"""
        try:
            start_col = self.start_col_var.get().upper()
            end_col = self.end_col_var.get().upper()
            start_row = int(self.start_row_var.get())
            end_row = int(self.end_row_var.get())

            # 验证列格式
            if not start_col.isalpha() or not end_col.isalpha():
                messagebox.showerror("错误", "列名必须是字母(A-Z)")
                return None

            # 验证行号
            if start_row < 1 or end_row < 1:
                messagebox.showerror("错误", "行号必须大于0")
                return None

            if start_row > end_row:
                messagebox.showerror("错误", "起始行不能大于结束行")
                return None

            start_col_idx = self._col_to_index(start_col)
            end_col_idx = self._col_to_index(end_col)

            if start_col_idx > end_col_idx:
                messagebox.showerror("错误", "起始列不能大于结束列")
                return None

            # 检查是否选择了排班模式
            selected_patterns = self._get_selected_patterns()
            if not selected_patterns:
                messagebox.showerror("错误", "请至少选择一种排班模式")
                return None

            return {
                'start_col': start_col,
                'end_col': end_col,
                'start_col_idx': start_col_idx,
                'end_col_idx': end_col_idx,
                'start_row': start_row,
                'end_row': end_row,
                'num_cols': end_col_idx - start_col_idx + 1,
                'num_rows': end_row - start_row + 1,
                'patterns': selected_patterns
            }

        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")
            return None

    def _generate_data_with_patterns(self, num_cols, num_rows, selected_patterns):
        """根据选定的模式生成数据"""
        all_data = []

        # 临时覆盖生成器的模式
        original_patterns = self.generator.SCHEDULE_PATTERNS.copy()

        # 只保留选中的模式
        filtered_patterns = {k: v for k, v in original_patterns.items() if k in selected_patterns}

        for _ in range(num_cols):
            col_data = []
            start_with_rest = self.allow_rest_start_var.get() and random.choice([True, False])

            # 从选中的模式中随机选择一个
            pattern_name = random.choice(list(filtered_patterns.keys()))
            work_days, rest_days = filtered_patterns[pattern_name]

            current_pos = 0

            # 如果从休息日开始
            if start_with_rest and rest_days > 0:
                rest_count = min(rest_days, num_rows - current_pos)
                for _ in range(rest_count):
                    col_data.append(None)
                    current_pos += 1

            # 循环生成工作和休息周期
            while current_pos < num_rows:
                # 工作周期
                work_count = min(work_days, num_rows - current_pos)
                for _ in range(work_count):
                    col_data.append(random.choice(self.generator.values))
                    current_pos += 1

                if current_pos >= num_rows:
                    break

                # 休息周期
                rest_count = min(rest_days, num_rows - current_pos)
                for _ in range(rest_count):
                    col_data.append(None)
                    current_pos += 1

            all_data.append(col_data)

        return all_data

    def _generate_and_save(self):
        """生成数据并保存到现有Excel文件"""
        params = self._validate_inputs()
        if not params:
            return

        if not self.selected_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return

        try:
            self.status_var.set("正在读取Excel文件...")
            self.root.update()

            # 加载工作簿
            wb = openpyxl.load_workbook(self.selected_file)

            # 如果有多个工作表，让用户选择
            sheet_names = wb.sheetnames
            if len(sheet_names) > 1:
                # 简单处理：使用第一个工作表，或者可以添加选择对话框
                ws = wb[sheet_names[0]]
            else:
                ws = wb.active

            self.status_var.set("正在生成排班数据...")
            self.root.update()

            # 生成数据
            all_data = self._generate_data_with_patterns(
                params['num_cols'],
                params['num_rows'],
                params['patterns']
            )

            # 写入数据
            self._write_data_to_worksheet(ws, all_data, params)

            # 保存文件
            self.status_var.set("正在保存文件...")
            self.root.update()

            wb.save(self.selected_file)

            self.status_var.set(f"数据已生成并保存到: {os.path.basename(self.selected_file)}")
            messagebox.showinfo("成功", f"排班数据已成功生成！\n\n文件: {self.selected_file}\n区域: {params['start_col']}:{params['end_col']}, 行{params['start_row']}-{params['end_row']}")

        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")
            messagebox.showerror("错误", f"处理文件时出错:\n{str(e)}")

    def _create_new_excel(self):
        """创建新的Excel文件并生成数据"""
        params = self._validate_inputs()
        if not params:
            return

        # 选择保存位置
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="排班数据.xlsx"
        )

        if not file_path:
            return

        try:
            self.status_var.set("正在创建新Excel文件...")
            self.root.update()

            # 创建新工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "排班数据"

            # 添加标题行
            ws['A1'] = "排班数据生成表"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"生成日期: {self._get_current_date()}"

            # 添加列标题（日期/员工标识）
            for col_idx in range(params['start_col_idx'], params['end_col_idx'] + 1):
                col_letter = self._index_to_col(col_idx)
                cell = ws.cell(row=params['start_row'] - 1, column=col_idx)
                cell.value = f"序列{col_letter}"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            self.status_var.set("正在生成排班数据...")
            self.root.update()

            # 生成数据
            all_data = self._generate_data_with_patterns(
                params['num_cols'],
                params['num_rows'],
                params['patterns']
            )

            # 写入数据
            self._write_data_to_worksheet(ws, all_data, params)

            # 调整列宽
            for col_idx in range(params['start_col_idx'], params['end_col_idx'] + 1):
                ws.column_dimensions[self._index_to_col(col_idx)].width = 8

            # 保存文件
            self.status_var.set("正在保存文件...")
            self.root.update()

            wb.save(file_path)

            self.selected_file = file_path
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)

            self.status_var.set(f"新文件已创建: {os.path.basename(file_path)}")
            messagebox.showinfo("成功", f"新Excel文件已创建并填充数据！\n\n文件: {file_path}")

        except Exception as e:
            self.status_var.set(f"错误: {str(e)}")
            messagebox.showerror("错误", f"创建文件时出错:\n{str(e)}")

    def _write_data_to_worksheet(self, ws, all_data, params):
        """将数据写入工作表"""
        # 定义样式
        work_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        rest_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        col_idx = params['start_col_idx']
        for col_data in all_data:
            for row_offset, value in enumerate(col_data):
                row = params['start_row'] + row_offset
                cell = ws.cell(row=row, column=col_idx)

                if value is None:
                    # 休息日
                    cell.value = "休"
                    cell.fill = rest_fill
                    cell.font = Font(color="666666", italic=True)
                else:
                    # 工作日
                    cell.value = value
                    cell.fill = work_fill
                    cell.font = Font(color="000000")

                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            col_idx += 1

    def _get_current_date(self):
        """获取当前日期字符串"""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d")


def main():
    """主函数"""
    root = tk.Tk()
    app = ScheduleGeneratorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
